from decimal import Decimal
import chromedriver_autoinstaller
import openpyxl
import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import zipfile
import json
import time
import os
import shutil
from datetime import datetime
import unittest
import math
import PySpice
import ltspice
import PyLTSpice
from PyLTSpice import SimCommander
from PySpice.Spice.Library import SpiceLibrary
from PySpice.Spice.Netlist import Circuit
from PySpice.Unit import *


class OpAmp(unittest.TestCase):

    def setUp(self):
        #driver instance
        options = Options()
        options.add_argument("--headless=new")
        chromedriver_autoinstaller.install()
        self.driver = webdriver.Chrome(options=options)
        with open(r'opAmp_OutputImpedance.json')as d:
            self.testData = json.load(d)['Variables'][0]

    def test_export(self):
        print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        print("        opAmp_OutputImpedance script is running...       ")
        print("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        driver = self.driver
        driver.maximize_window()

        # based on the json file the script will select With Load or Without Load Design.
        if (self.testData['load'] == 'Yes'):
            driver.get(self.testData['URL_with_load'])

            #Accept Cookies
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#noise-spinner")))
            WebDriverWait(driver, 10).until(EC.invisibility_of_element((By.CSS_SELECTOR, "#noise-spinner")))
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, "body.ember-application:nth-child(2) div.consent-dialog:nth-child(1) div.modal.fade.in.show "
                                "div.modal-dialog div.modal-content div.modal-body div.short-description > a.btn.btn-success:nth-child(2)"))).click()
        else:
            driver.get(self.testData['URL_without_load'])

            #Accept Cookies
            WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#noise-spinner")))
            WebDriverWait(driver, 10).until(EC.invisibility_of_element((By.CSS_SELECTOR, "#noise-spinner")))
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((
                By.CSS_SELECTOR, "body.ember-application:nth-child(2) div.consent-dialog:nth-child(1) div.modal.fade.in.show "
                                "div.modal-dialog div.modal-content div.modal-body div.short-description > a.btn.btn-success:nth-child(2)"))).click()

        gain = self.testData['gain']
        device = self.testData['device']
        R2 = self.testData['R2']
        C2VALUE = self.testData['C2VALUE']

        #Run the simulation in Nimble
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((
            By.XPATH, "//body/div[@id='base-container']/div[@id='main-content-container']/div[@id='application-view']/div[@id='build-signal-chain-tab-content']"
            "/div[@id='adi-signal-chain-row']/div[@id='analog-signal-chain-group']/div[@id='signal-chain-drop-area']/table[1]/tr[1]/td[1]/div[1]/div[2]/div[2]/div[1]/*[1]" ))).click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#amp-gain-input'))).click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#amp-gain-input'))).send_keys(Keys.CONTROL + "a")
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#amp-gain-input'))).send_keys(Keys.DELETE)
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#amp-gain-input'))).send_keys(gain)
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((
            By.CSS_SELECTOR, "#text6747-2-6 > tspan.schematic-edit-icon.schematic-part-edit-selection-link.schematic-edit-selection-link" ))).click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#filter-0'))).click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#filter-0'))).send_keys(device)
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//body/div[@id='base-container']/div[@id='main-content-container']/div[@id='application-view']/div[@id='config-signal-chain-item-modal']"
                          "/div[1]/div[1]/div[1]/div[2]/div[1]/div[4]/div[1]/div[1]/div[1]/div[2]/div[1]/div[2]/div[4]/div[3]/div[1]/div[1]/div[1]"))).click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((
            By.CSS_SELECTOR, 'body.ember-application.modal-open:nth-child(2) div.adi-modal.modal-fills-window.modal-hide-scroll:nth-child(5) div.modal.fade.show.d-block:nth-child(1) '
            'div.modal-dialog div.modal-content div.modal-body div.configure-amp.configure-signal-chain-item div.adi-modal.modal-fills-window:nth-child(5) '
            'div.modal.fade.show.d-block:nth-child(1) div.modal-dialog div.modal-content div.modal-footer div.button-row > button.btn.btn-primary:nth-child(1)' ))).click()

        # Dictionary converting kilo, Mega
        d = {'k': 1000, 'M': 1000000, 'f': 1e-15, 'p': 1e-12, 'n': 1e-9, 'u': 1e-6}
        def text_to_num(text):
            if text[-1] in d:
                num, magnitude = text[:-1], text[-1]
                return float(num) * d[magnitude]
            else:
                return Decimal(text)
        new_rvalue = text_to_num(R2)
        new_c2value = text_to_num(C2VALUE)
  
        #This function gets the Slider Value to be passed by javascript command     
        def value_to_position(value, limit1, limit2):
            minpos = 1
            maxpos = 10000
            minval = math.log(limit1)
            maxval = math.log(limit2)
            scale = (maxval - minval) / (maxpos - minpos)

            if value <= 0: 
                return minpos 
            else: 
                position = minpos + (math.log(value) - minval) / scale 
                return position
            
        rposition = value_to_position(new_rvalue, 10, 10000000)
        c2position = value_to_position(new_c2value, 1e-15, 1e-6)

        if (float(gain)!= 1):
            driver.execute_script(f"document.querySelector('#rscale-slider').value = {rposition}; document.querySelector('#rscale-slider').dispatchEvent(new Event('input'));")
            driver.execute_script(f"document.querySelector('#c2-slider').value = {c2position}; document.querySelector('#c2-slider').dispatchEvent(new Event('input'));")
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            print ("                    Slider values set!                   ")
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        else:
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
            print ("                     No Slider values                    ")
            print ("+++++++++++++++++++++++++++++++++++++++++++++++++++++++++")
        time.sleep(2)

        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//*[@id=\"config-signal-chain-item-modal\"]/div[1]/div/div/div[3]/div/button[1]"))).click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#next-steps-tab"))).click()
        WebDriverWait(driver, 5).until(EC.presence_of_element_located((
            By.XPATH, "//body/div[@id='base-container']/div[@id='main-content-container']/div[@id='application-view']/"
                      "div[@id='next-steps-tab-content']/div[@id='next-steps-container']/div[2]/div[1]"))).click()
        time.sleep(5)

        #This script is extracting files directly to project folder
        now = datetime.now()
        day = str(now.day)
        current_date = now.strftime(f"%B {day}, %Y")
        downloads_path = self.testData['downloads_path']
        project_path = self.testData['project_location']
        file_path = downloads_path + 'Full Data Export - ' + current_date + '.zip'
        with zipfile.ZipFile(file_path) as zip_ref:
            new_path = project_path + '\\' + device + ' extracted files'
            zip_ref.extractall(new_path)
        print("Files were extracted to project folder")

        #Deletes the zip file after extracting
        if os.path.exists(file_path):
            os.remove(file_path)
        else:
            print("Zip file does not exist")  

        # Running the simulation in LTSpice          
        file_path = project_path + '\\' + device + ' extracted files' + '\\' + 'Ltspice Schematics'

        # netlists are created
        LTC = SimCommander(file_path + "\\AC_Simulation.asc")

        #changing netlist file into txt file
        old_path = (file_path + "\\AC_Simulation.net")
        new_path = (file_path + "\\AC_Simulation_Result.txt")
        
        dir_path, filename = os.path.split(old_path)
        name, ext = os.path.splitext(filename)
        new_name, new_ext = os.path.splitext(new_path)
        new_filename = new_name + new_ext
        new_path = os.path.join(dir_path, new_filename)
        shutil.copy(old_path, new_path)

        #making changes in the txt file
        with open(new_path, "r") as f:
            lines = f.readlines()
        del lines[1]
        lines[1:3] = ["R1 N002 0 1\n", "C1 N002 0 1f\n"]
        lines.insert(8, "I1 0 out 0 AC 1\n")
        line_to_modify = 3  
        position_to_modify = 9  
        new_character = "2"
        lines[line_to_modify] = lines[line_to_modify][:position_to_modify] + new_character + lines[line_to_modify][position_to_modify + 1:]

        #extract values of VDD-1 and VSS-1 and store them in variables
        line7 = lines[6].rstrip()  # remove trailing newline character
        num1_str = line7.split()[-1]  # extract last space-separated element of line
        sym1 = float(num1_str)  # convert string to float

        line8 = lines[7].rstrip()  # remove trailing newline character
        num2_str = line8.split()[-1]  # extract last space-separated element of line
        sym2 = float(num2_str)  # convert string to float

        #making numbers symmetrical
        def symmetrical(num1, num2):
            avg = (num1 + num2) / 2
            sym_num1 = 2 * avg - num1
            sym_num2 = 2 * avg - num2
            if sym_num1 == sym_num2:
                return sym_num1, -sym_num2
            else:
                return (num1 + num2) / 2, -(num1 + num2) / 2
            
        num1 = sym1
        num2 = sym2
        sym_num1, sym_num2 = symmetrical(num1, num2)

        line7_new = line7.replace(num1_str, str(sym_num1))  # replace last number with sym_num1
        lines[6] = line7_new + "\n"  # add newline character back and update the list

        line8_new = lines[7].rsplit(' ', 1)[0] + f' {sym_num2}\n'
        lines[7] = line8_new

        with open(new_path, "w") as f:
            f.writelines(lines)

        #changing txt file back into netlist file
        old_path1 = (file_path + "\\AC_Simulation_Result.txt")
        new_path1 = (file_path + "\\AC_Simulation_Result.net")
        
        dir_path, filename = os.path.split(old_path1)
        name, ext = os.path.splitext(filename)
        new_name, new_ext = os.path.splitext(new_path1)
        new_filename = new_name + new_ext
        new_path = os.path.join(dir_path, new_filename)
        shutil.copy(old_path1, new_path1)

        # Parse the LTSpice raw file
        l = ltspice.Ltspice(file_path + "\\AC_Simulation_Result.raw")
        l.parse()

        # Get the V(out) trace data
        freq = l.get_frequency()
        Vout = l.get_data('V(out)')

        # Create a DataFrame with the frequency and V(onoise) data
        data = {'Frequency (Hz)': freq, 'V(out)': Vout}
        df = pd.DataFrame(data)

        # Export the DataFrame to an Excel file
        ltspice_output_path = (project_path + '\\' + device + '_Output_Impedance.xlsx')
        df.to_excel(ltspice_output_path, index=False, engine='openpyxl')
      
        # Converting the Amplifier - Input Referred Noise.csv to .xlsx
        path_file = pd.read_csv(project_path + '\\' + device + ' extracted files' + '\\' + 'Raw Data' + '\\' + 'Individual Stage Data' + '\\' + 'Amplifier' + '\\' + 'Amplifier - Input and Output Impedance.csv')
        nimble_output_path = project_path + '\\' + device + ' Amplifier - Input and Output Impedance.xlsx'
        path_file.to_excel(nimble_output_path, index=None, header=True)

        # Deleting the extra collumns
        file = openpyxl.load_workbook(nimble_output_path)
        sheet_obj = file.active
        sheet_obj.delete_cols(2)
        sheet_obj.delete_cols(2)
        sheet_obj.delete_cols(3)
        file.save(nimble_output_path)

        # Getting the data from Output_Impedance.xlsx to Amplifier - Input Referred Noise.xlsx in a new sheet
        wb1 = openpyxl.load_workbook(filename=ltspice_output_path)
        ws1 = wb1.worksheets[0]
        wb2 = openpyxl.load_workbook(filename=nimble_output_path)
        ws2 = wb2.create_sheet(ws1.title)

        for row in ws1:
            for cell in row:
                ws2[cell.coordinate].value = cell.value

        wb2.save(nimble_output_path)
        


        

            

        #WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '' ))).click()








    def tearDown(self):
        self.driver.quit()

if __name__ == '__main__':
    unittest.main()        
